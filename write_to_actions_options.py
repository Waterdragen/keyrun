import json
from openpyxl import load_workbook

def main():
    wb = load_workbook("action_options.xlsx", data_only=True, read_only=True)
    ws = wb.active
    header_cols = ('B', 'C', 'D', 'E', 'F')
    actions_config: dict = {}
    for r in range(2, 100):
        action = ws[f"A{r}"].value
        if action:
            actions_config[action] = {ws[f"{header}1"].value: ws[f"{header}{r}"].value for header in header_cols}
        else:
            break
    with open("action_options.json", "w") as f:
        json.dump(actions_config, f, indent=4)


if __name__ == '__main__':
    main()
